# archivo: grupo_operativo.py
from fastapi import APIRouter, Depends, HTTPException, Query
from starlette.responses import StreamingResponse
import pyodbc
import io
import csv

from database import get_connection
from security import get_current_user

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(
    prefix="/grupo-operativo",
    tags=["GrupoOperativo"],
    dependencies=[Depends(get_current_user)],
)


def _row_to_dict(r):
    return {
        "PKID": r.PKID,
        "IDGrupoOperativo": r.IDGrupoOperativo,
        "GrupoOperativo": r.GrupoOperativo,
        "PKIDSituacionRegistro": r.PKIDSituacionRegistro,
        "SituacionRegistro": r.SituacionRegistro,
    }


def _validate_payload(p: dict):
    # IDGrupoOperativo obligatorio
    if p.get("IDGrupoOperativo") in (None, ""):
        raise HTTPException(status_code=422, detail="IDGrupoOperativo es obligatorio")
    try:
        p["IDGrupoOperativo"] = int(p["IDGrupoOperativo"])
    except Exception:
        raise HTTPException(status_code=422, detail="IDGrupoOperativo debe ser entero")

    nombre = (p.get("GrupoOperativo") or "").strip()
    if not nombre:
        raise HTTPException(status_code=422, detail="GrupoOperativo es obligatorio")
    if len(nombre) > 50:
        raise HTTPException(status_code=422, detail="GrupoOperativo máximo 50 caracteres")
    p["GrupoOperativo"] = nombre

    if p.get("PKIDSituacionRegistro") in (None, ""):
        raise HTTPException(status_code=422, detail="PKIDSituacionRegistro es obligatorio")
    try:
        p["PKIDSituacionRegistro"] = int(p["PKIDSituacionRegistro"])
    except Exception:
        raise HTTPException(status_code=422, detail="PKIDSituacionRegistro debe ser entero")


def _build_where(params, idgrupo=None, nombre=None, situacion_id=None):
    where = " WHERE 1=1"
    if idgrupo not in (None, ""):
        # si es número exacto
        if str(idgrupo).isdigit():
            where += " AND g.IDGrupoOperativo = ?"
            params.append(int(idgrupo))
        else:
            where += " AND CAST(g.IDGrupoOperativo AS VARCHAR(50)) LIKE ?"
            params.append(f"%{idgrupo.strip()}%")
    if nombre:
        where += " AND g.GrupoOperativo LIKE ?"
        params.append(f"%{nombre.strip()}%")
    if situacion_id not in (None, ""):
        where += " AND g.PKIDSituacionRegistro = ?"
        params.append(int(situacion_id))
    return where


@router.get("/")
def list_grupo_operativo(
    idgrupo: str | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
):
    """
    Lista con paginación y filtros.
    """
    try:
        conn = get_connection()
        cur = conn.cursor()

        params_total = []
        where_total = _build_where(params_total, idgrupo, nombre, situacion_id)
        cur.execute(f"""
            SELECT COUNT(1) AS total
            FROM GrupoOperativo g
            LEFT JOIN SituacionRegistro s ON g.PKIDSituacionRegistro = s.PKID
            {where_total}
        """, params_total)
        total = int(cur.fetchone().total)

        offset = (page - 1) * page_size
        params = []
        where = _build_where(params, idgrupo, nombre, situacion_id)
        cur.execute(f"""
            SELECT g.PKID, g.IDGrupoOperativo, g.GrupoOperativo,
                   g.PKIDSituacionRegistro, s.SituacionRegistro
            FROM GrupoOperativo g
            LEFT JOIN SituacionRegistro s ON g.PKIDSituacionRegistro = s.PKID
            {where}
            ORDER BY g.IDGrupoOperativo ASC, g.GrupoOperativo ASC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, page_size])
        items = [_row_to_dict(r) for r in cur.fetchall()]
        return {"items": items, "total": total}
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/")
def create_grupo_operativo(payload: dict):
    _validate_payload(payload)
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO GrupoOperativo
              (IDGrupoOperativo, GrupoOperativo, PKIDSituacionRegistro)
            VALUES (?,?,?)
        """, (
            payload["IDGrupoOperativo"],
            payload["GrupoOperativo"],
            payload["PKIDSituacionRegistro"],
        ))
        conn.commit()
        return {"detail": "Creado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="IDGrupoOperativo duplicado o FK inválida.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{pkid}")
def update_grupo_operativo(pkid: int, payload: dict):
    _validate_payload(payload)
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM GrupoOperativo WHERE PKID = ?", (pkid,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="No encontrado.")

        cur.execute("""
            UPDATE GrupoOperativo
               SET IDGrupoOperativo = ?,
                   GrupoOperativo = ?,
                   PKIDSituacionRegistro = ?
             WHERE PKID = ?
        """, (
            payload["IDGrupoOperativo"],
            payload["GrupoOperativo"],
            payload["PKIDSituacionRegistro"],
            pkid,
        ))
        conn.commit()
        return {"detail": "Actualizado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="IDGrupoOperativo duplicado o FK inválida.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{pkid}")
def delete_grupo_operativo(pkid: int):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM GrupoOperativo WHERE PKID = ?", (pkid,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado.")
        conn.commit()
        return {"detail": "Eliminado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="No se puede eliminar, está en uso.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export")
def export_csv(
    idgrupo: str | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    """
    Exporta CSV con filtros.
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        where = _build_where(params, idgrupo, nombre, situacion_id)
        cur.execute(f"""
            SELECT g.IDGrupoOperativo, g.GrupoOperativo, s.SituacionRegistro
            FROM GrupoOperativo g
            LEFT JOIN SituacionRegistro s ON g.PKIDSituacionRegistro = s.PKID
            {where}
            ORDER BY g.IDGrupoOperativo ASC, g.GrupoOperativo ASC
        """, params)
        rows = cur.fetchall()

        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(["IDGrupoOperativo", "GrupoOperativo", "Situacion"])
        for r in rows:
            w.writerow([r.IDGrupoOperativo, r.GrupoOperativo, r.SituacionRegistro or ""])

        bio = io.BytesIO(si.getvalue().encode("utf-8-sig"))
        headers = {"Content-Disposition": 'attachment; filename="grupo_operativo.csv"'}
        return StreamingResponse(bio, media_type="text/csv", headers=headers)
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-xlsx")
def export_xlsx(
    idgrupo: str | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    """
    Exporta XLSX con filtros.
    """
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        where = _build_where(params, idgrupo, nombre, situacion_id)
        cur.execute(f"""
            SELECT g.IDGrupoOperativo, g.GrupoOperativo, s.SituacionRegistro
            FROM GrupoOperativo g
            LEFT JOIN SituacionRegistro s ON g.PKIDSituacionRegistro = s.PKID
            {where}
            ORDER BY g.IDGrupoOperativo ASC, g.GrupoOperativo ASC
        """, params)
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "GrupoOperativo"
        headers = ["IDGrupoOperativo", "GrupoOperativo", "Situación"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="DDDDDD")

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for r in rows:
            ws.append([r.IDGrupoOperativo, r.GrupoOperativo, r.SituacionRegistro or ""])

        from openpyxl.utils import get_column_letter
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                max_len = max(max_len, len(str(val)) if val else 0)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="grupo_operativo.xlsx"'}
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
