# archivo: familia.py
from fastapi import APIRouter, Depends, HTTPException, Query
from starlette.responses import StreamingResponse
import pyodbc
import io
import csv

from database import get_connection
from security import get_current_user

# NUEVO: XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(
    prefix="/familia",
    tags=["Familia"],
    dependencies=[Depends(get_current_user)],
)

def _row_to_dict(r):
    return {
        "PKID": r.PKID,
        "IDFamilia": r.IDFamilia,
        "Familia": r.Familia,
        "AcumulativaCheck": bool(r.AcumulativaCheck),
        "FijaCheck": bool(r.FijaCheck),
        "CompuestaCheck": bool(r.CompuestaCheck),
        "PKIDSituacionRegistro": r.PKIDSituacionRegistro,
        "SituacionRegistro": r.SituacionRegistro,
    }

def _build_where(params, id_familia, nombre, situacion_id):
    sql_where = " WHERE 1=1"
    if id_familia is not None:
        sql_where += " AND f.IDFamilia = ?"
        params.append(id_familia)
    if nombre:
        sql_where += " AND f.Familia LIKE ?"
        params.append(f"%{nombre}%")
    if situacion_id is not None:
        sql_where += " AND f.PKIDSituacionRegistro = ?"
        params.append(situacion_id)
    return sql_where

@router.get("/")
def list_familias(
    id_familia: int | None = Query(default=None, description="Filtra por IDFamilia exacto"),
    nombre: str | None = Query(default=None, description="Filtra por Familia (contiene)"),
    situacion_id: int | None = Query(default=None, description="Filtra por PKIDSituacionRegistro exacto"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
):
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Total
        params_total = []
        where_total = _build_where(params_total, id_familia, nombre, situacion_id)
        sql_total = f"""
            SELECT COUNT(1) AS total
            FROM Familia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where_total}
        """
        cur.execute(sql_total, params_total)
        total_row = cur.fetchone()
        total = int(total_row.total) if total_row else 0

        # Página
        offset = (page - 1) * page_size
        params = []
        where = _build_where(params, id_familia, nombre, situacion_id)
        sql = f"""
            SELECT f.PKID, f.IDFamilia, f.Familia,
                   f.AcumulativaCheck, f.FijaCheck, f.CompuestaCheck,
                   f.PKIDSituacionRegistro, sr.SituacionRegistro
            FROM Familia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY f.IDFamilia
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        params += [offset, page_size]
        cur.execute(sql, params)
        items = [_row_to_dict(r) for r in cur.fetchall()]

        return {"items": items, "total": total}
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/")
def create_familia(payload: dict):
    required = [
        "IDFamilia", "Familia",
        "AcumulativaCheck", "FijaCheck", "CompuestaCheck",
        "PKIDSituacionRegistro"
    ]
    for k in required:
        if payload.get(k) in (None, ""):
            raise HTTPException(status_code=422, detail=f"Falta campo requerido: {k}")
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Unicidad IDFamilia
        cur.execute("SELECT 1 FROM Familia WHERE IDFamilia = ?", (payload["IDFamilia"],))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="IDFamilia ya existe.")

        cur.execute("""
            INSERT INTO Familia (
                IDFamilia, Familia,
                AcumulativaCheck, FijaCheck, CompuestaCheck,
                PKIDSituacionRegistro
            )
            VALUES (?,?,?,?,?,?)
        """, (
            int(payload["IDFamilia"]),
            str(payload["Familia"]).strip(),
            1 if bool(payload["AcumulativaCheck"]) else 0,
            1 if bool(payload["FijaCheck"]) else 0,
            1 if bool(payload["CompuestaCheck"]) else 0,
            int(payload["PKIDSituacionRegistro"]),
        ))
        conn.commit()
        return {"detail": "Creado"}
    except HTTPException:
        raise
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad / llave duplicada / FK.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{pkid}")
def update_familia(pkid: int, payload: dict):
    try:
        conn = get_connection()
        cur = conn.cursor()

        # Existe
        cur.execute("SELECT PKID, IDFamilia FROM Familia WHERE PKID = ?", (pkid,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Familia no encontrada.")

        # Unicidad si cambia IDFamilia
        _idfam = int(payload.get("IDFamilia", row.IDFamilia))
        cur.execute("SELECT 1 FROM Familia WHERE IDFamilia = ? AND PKID <> ?", (_idfam, pkid))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="IDFamilia ya existe para otro registro.")

        cur.execute("""
            UPDATE Familia
               SET IDFamilia = ?,
                   Familia = ?,
                   AcumulativaCheck = ?,
                   FijaCheck = ?,
                   CompuestaCheck = ?,
                   PKIDSituacionRegistro = ?
             WHERE PKID = ?
        """, (
            _idfam,
            str(payload["Familia"]).strip(),
            1 if bool(payload["AcumulativaCheck"]) else 0,
            1 if bool(payload["FijaCheck"]) else 0,
            1 if bool(payload["CompuestaCheck"]) else 0,
            int(payload["PKIDSituacionRegistro"]),
            pkid,
        ))
        conn.commit()
        return {"detail": "Actualizado"}
    except HTTPException:
        raise
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad / FK / llave duplicada.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{pkid}")
def delete_familia(pkid: int):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM Familia WHERE PKID = ?", (pkid,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Familia no encontrada.")
        conn.commit()
        return {"detail": "Eliminado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="No se puede eliminar: hay referencias.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export")
def export_familias_csv(
    id_familia: int | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    """
    Exporta los datos filtrados a CSV (Excel-compatible).
    """
    try:
        conn = get_connection()
        cur = conn.cursor()

        params = []
        where = _build_where(params, id_familia, nombre, situacion_id)
        sql = f"""
            SELECT f.IDFamilia, f.Familia,
                   f.AcumulativaCheck, f.FijaCheck, f.CompuestaCheck,
                   sr.SituacionRegistro
            FROM Familia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY f.IDFamilia
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        # CSV en memoria
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(["IDFamilia", "Familia", "Acumulativa", "Fija", "Compuesta", "Situacion"])
        for r in rows:
            writer.writerow([
                r.IDFamilia,
                r.Familia,
                1 if r.AcumulativaCheck else 0,
                1 if r.FijaCheck else 0,
                1 if r.CompuestaCheck else 0,
                r.SituacionRegistro,
            ])
        out = io.BytesIO(si.getvalue().encode("utf-8-sig"))
        headers = {
            "Content-Disposition": 'attachment; filename="familias.csv"'
        }
        return StreamingResponse(out, media_type="text/csv", headers=headers)
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-xlsx")
def export_familias_xlsx(
    id_familia: int | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    """
    Exporta los datos filtrados a XLSX (Excel nativo).
    """
    try:
        conn = get_connection()
        cur = conn.cursor()

        params = []
        where = _build_where(params, id_familia, nombre, situacion_id)
        sql = f"""
            SELECT f.IDFamilia, f.Familia,
                   f.AcumulativaCheck, f.FijaCheck, f.CompuestaCheck,
                   sr.SituacionRegistro
            FROM Familia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY f.IDFamilia
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        # Construir XLSX con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Familias"

        headers = ["IDFamilia", "Familia", "Acumulativa", "Fija", "Compuesta", "Situación"]
        ws.append(headers)

        # Estilos de encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")  # Indigo
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")

        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Data
        for r in rows:
            ws.append([
                r.IDFamilia,
                r.Familia,
                "Sí" if r.AcumulativaCheck else "No",
                "Sí" if r.FijaCheck else "No",
                "Sí" if r.CompuestaCheck else "No",
                r.SituacionRegistro,
            ])

        # Auto ancho columnas
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                ln = len(str(val)) if val is not None else 0
                max_len = max(max_len, ln)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"

        # Output
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        headers_resp = {
            "Content-Disposition": 'attachment; filename="familias.xlsx"'
        }
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_resp
        )
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
