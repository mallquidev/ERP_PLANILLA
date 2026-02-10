# archivo: familia_remuneracion_variable.py
from fastapi import APIRouter, Depends, HTTPException, Query
from starlette.responses import StreamingResponse
import pyodbc
import io
import csv

from database import get_connection
from security import get_current_user

# XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(
    prefix="/familia-remuneracion-variable",
    tags=["FamiliaRemuneracionVariable"],
    dependencies=[Depends(get_current_user)],
)

# --------------------- helpers ---------------------

def _row_to_dict(r):
    return {
        "PKID": r.PKID,
        "IDFamiliaRemuneracionVariable": r.IDFamiliaRemuneracionVariable,
        "FamiliaRemuneracionVariable": r.FamiliaRemuneracionVariable,
        "PKIDSituacionRegistro": r.PKIDSituacionRegistro,
        "SituacionRegistro": r.SituacionRegistro,
    }

def _build_where(params, id_codigo, nombre, situacion_id):
    where = " WHERE 1=1"
    if id_codigo is not None and id_codigo != "":
        where += " AND frv.IDFamiliaRemuneracionVariable = ?"
        params.append(int(id_codigo))
    if nombre:
        where += " AND frv.FamiliaRemuneracionVariable LIKE ?"
        params.append(f"%{nombre}%")
    if situacion_id is not None and situacion_id != "":
        where += " AND frv.PKIDSituacionRegistro = ?"
        params.append(int(situacion_id))
    return where

def _validate_payload(payload: dict, updating: bool = False):
    # IDFamiliaRemuneracionVariable requerido al crear; al actualizar puede venir o no (si cambia)
    if not updating:
        if payload.get("IDFamiliaRemuneracionVariable") in (None, ""):
            raise HTTPException(status_code=422, detail="Falta campo requerido: IDFamiliaRemuneracionVariable")
    if "IDFamiliaRemuneracionVariable" in payload and payload["IDFamiliaRemuneracionVariable"] not in (None, ""):
        try:
            v = int(payload["IDFamiliaRemuneracionVariable"])
            if v <= 0:
                raise ValueError()
        except Exception:
            raise HTTPException(status_code=422, detail="IDFamiliaRemuneracionVariable debe ser entero positivo")

    nombre = payload.get("FamiliaRemuneracionVariable")
    if nombre is None or str(nombre).strip() == "":
        raise HTTPException(status_code=422, detail="Falta campo requerido: FamiliaRemuneracionVariable")
    if len(str(nombre).strip()) > 50:
        raise HTTPException(status_code=422, detail="FamiliaRemuneracionVariable debe tener como máximo 50 caracteres")

    if payload.get("PKIDSituacionRegistro") not in (None, ""):
        try:
            int(payload["PKIDSituacionRegistro"])
        except Exception:
            raise HTTPException(status_code=422, detail="PKIDSituacionRegistro debe ser entero")

# --------------------- list ---------------------

@router.get("/")
def list_familia_rem_var(
    id_codigo: int | None = Query(default=None, description="Filtra por IDFamiliaRemuneracionVariable exacto"),
    nombre: str | None = Query(default=None, description="Filtra por nombre (contiene)"),
    situacion_id: int | None = Query(default=None, description="Filtra por PKIDSituacionRegistro exacto"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
):
    try:
        conn = get_connection()
        cur = conn.cursor()

        # total
        params_total = []
        where_total = _build_where(params_total, id_codigo, nombre, situacion_id)
        sql_total = f"""
            SELECT COUNT(1) AS total
            FROM FamiliaRemuneracionVariable frv
            LEFT JOIN SituacionRegistro sr ON frv.PKIDSituacionRegistro = sr.PKID
            {where_total}
        """
        cur.execute(sql_total, params_total)
        total = int(cur.fetchone().total)

        # page
        offset = (page - 1) * page_size
        params = []
        where = _build_where(params, id_codigo, nombre, situacion_id)
        sql = f"""
            SELECT frv.PKID, frv.IDFamiliaRemuneracionVariable, frv.FamiliaRemuneracionVariable,
                   frv.PKIDSituacionRegistro, sr.SituacionRegistro
            FROM FamiliaRemuneracionVariable frv
            LEFT JOIN SituacionRegistro sr ON frv.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY frv.IDFamiliaRemuneracionVariable
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        params += [offset, page_size]
        cur.execute(sql, params)
        items = [_row_to_dict(r) for r in cur.fetchall()]
        return {"items": items, "total": total}
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- create ---------------------

@router.post("/")
def create_familia_rem_var(payload: dict):
    _validate_payload(payload, updating=False)
    try:
        conn = get_connection()
        cur = conn.cursor()

        # unicidad por IDFamiliaRemuneracionVariable
        cur.execute(
            "SELECT 1 FROM FamiliaRemuneracionVariable WHERE IDFamiliaRemuneracionVariable = ?",
            (int(payload["IDFamiliaRemuneracionVariable"]),),
        )
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="IDFamiliaRemuneracionVariable ya existe.")

        cur.execute("""
            INSERT INTO FamiliaRemuneracionVariable
              (IDFamiliaRemuneracionVariable, FamiliaRemuneracionVariable, PKIDSituacionRegistro)
            VALUES (?,?,?)
        """, (
            int(payload["IDFamiliaRemuneracionVariable"]),
            str(payload["FamiliaRemuneracionVariable"]).strip(),
            int(payload["PKIDSituacionRegistro"]) if payload.get("PKIDSituacionRegistro") not in (None, "",) else None,
        ))
        conn.commit()
        return {"detail": "Creado"}
    except HTTPException:
        raise
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad / llave duplicada / FK.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- update ---------------------

@router.put("/{pkid}")
def update_familia_rem_var(pkid: int, payload: dict):
    _validate_payload(payload, updating=True)
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT PKID, IDFamiliaRemuneracionVariable FROM FamiliaRemuneracionVariable WHERE PKID = ?", (pkid,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="No encontrado.")

        new_id = int(payload.get("IDFamiliaRemuneracionVariable", row.IDFamiliaRemuneracionVariable))
        if new_id != row.IDFamiliaRemuneracionVariable:
            cur.execute(
                "SELECT 1 FROM FamiliaRemuneracionVariable WHERE IDFamiliaRemuneracionVariable = ? AND PKID <> ?",
                (new_id, pkid),
            )
            if cur.fetchone():
                raise HTTPException(status_code=409, detail="IDFamiliaRemuneracionVariable ya existe para otro registro.")

        cur.execute("""
            UPDATE FamiliaRemuneracionVariable
               SET IDFamiliaRemuneracionVariable = ?,
                   FamiliaRemuneracionVariable = ?,
                   PKIDSituacionRegistro = ?
             WHERE PKID = ?
        """, (
            new_id,
            str(payload["FamiliaRemuneracionVariable"]).strip(),
            int(payload["PKIDSituacionRegistro"]) if payload.get("PKIDSituacionRegistro") not in (None, "",) else None,
            pkid,
        ))
        conn.commit()
        return {"detail": "Actualizado"}
    except HTTPException:
        raise
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad / FK.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- delete ---------------------

@router.delete("/{pkid}")
def delete_familia_rem_var(pkid: int):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM FamiliaRemuneracionVariable WHERE PKID = ?", (pkid,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado.")
        conn.commit()
        return {"detail": "Eliminado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="No se puede eliminar: hay referencias.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- export CSV ---------------------

@router.get("/export")
def export_csv(
    id_codigo: int | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    try:
        conn = get_connection()
        cur = conn.cursor()

        params = []
        where = _build_where(params, id_codigo, nombre, situacion_id)
        sql = f"""
            SELECT frv.IDFamiliaRemuneracionVariable, frv.FamiliaRemuneracionVariable, sr.SituacionRegistro
            FROM FamiliaRemuneracionVariable frv
            LEFT JOIN SituacionRegistro sr ON frv.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY frv.IDFamiliaRemuneracionVariable
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(["ID", "Nombre", "Situación"])
        for r in rows:
            w.writerow([r.IDFamiliaRemuneracionVariable, r.FamiliaRemuneracionVariable, r.SituacionRegistro])

        out = io.BytesIO(si.getvalue().encode("utf-8-sig"))
        headers = {"Content-Disposition": 'attachment; filename="familia_remuneracion_variable.csv"'}
        return StreamingResponse(out, media_type="text/csv", headers=headers)
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- export XLSX ---------------------

@router.get("/export-xlsx")
def export_xlsx(
    id_codigo: int | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    try:
        conn = get_connection()
        cur = conn.cursor()

        params = []
        where = _build_where(params, id_codigo, nombre, situacion_id)
        sql = f"""
            SELECT frv.IDFamiliaRemuneracionVariable, frv.FamiliaRemuneracionVariable, sr.SituacionRegistro
            FROM FamiliaRemuneracionVariable frv
            LEFT JOIN SituacionRegistro sr ON frv.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY frv.IDFamiliaRemuneracionVariable
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "FamiliaRemVar"
        headers = ["ID", "Nombre", "Situación"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for r in rows:
            ws.append([r.IDFamiliaRemuneracionVariable, r.FamiliaRemuneracionVariable, r.SituacionRegistro])

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        ws.freeze_panes = "A2"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="familia_remuneracion_variable.xlsx"'}
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
