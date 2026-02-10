# archivo: frecuencia.py
from fastapi import APIRouter, Depends, HTTPException, Query
from starlette.responses import StreamingResponse
import pyodbc
import io
import csv

from database import get_connection
from security import get_current_user

# XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(
    prefix="/frecuencia",
    tags=["Frecuencia"],
    dependencies=[Depends(get_current_user)],
)

# --------------------- helpers ---------------------

def _row_to_dict(r):
    return {
        "PKID": r.PKID,
        "IDFrecuencia": r.IDFrecuencia,
        "Frecuencia": r.Frecuencia,
        "NumeroDias": r.NumeroDias,
        "NumeroHoras": r.NumeroHoras,
        "PKIDSituacionRegistro": r.PKIDSituacionRegistro,
        "SituacionRegistro": r.SituacionRegistro,
    }

def _validate_payload(p: dict):
    # IDFrecuencia: opcional (int)
    if p.get("IDFrecuencia") not in (None, ""):
        try:
            p["IDFrecuencia"] = int(p["IDFrecuencia"])
        except Exception:
            raise HTTPException(status_code=422, detail="IDFrecuencia debe ser entero")

    # Frecuencia: requerido
    name = (p.get("Frecuencia") or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Falta campo requerido: Frecuencia")
    if len(name) > 50:
        raise HTTPException(status_code=422, detail="Frecuencia: máximo 50 caracteres")
    p["Frecuencia"] = name

    # NumeroDias: requerido int >= 0
    try:
        nd = int(p.get("NumeroDias"))
    except Exception:
        raise HTTPException(status_code=422, detail="NumeroDias debe ser entero")
    if nd < 0:
        raise HTTPException(status_code=422, detail="NumeroDias no puede ser negativo")
    p["NumeroDias"] = nd

    # NumeroHoras: requerido int >= 0
    try:
        nh = int(p.get("NumeroHoras"))
    except Exception:
        raise HTTPException(status_code=422, detail="NumeroHoras debe ser entero")
    if nh < 0:
        raise HTTPException(status_code=422, detail="NumeroHoras no puede ser negativo")
    p["NumeroHoras"] = nh

    # PKIDSituacionRegistro: requerido int
    if p.get("PKIDSituacionRegistro") in (None, ""):
        raise HTTPException(status_code=422, detail="Falta campo requerido: PKIDSituacionRegistro")
    try:
        p["PKIDSituacionRegistro"] = int(p["PKIDSituacionRegistro"])
    except Exception:
        raise HTTPException(status_code=422, detail="PKIDSituacionRegistro debe ser entero")

def _build_where(params, idfrecuencia: str | None, nombre: str | None, situacion_id: int | None):
    where = " WHERE 1=1"
    if idfrecuencia not in (None, ""):
        # busca exacto o por coincidencia si viene texto
        if str(idfrecuencia).isdigit():
            where += " AND f.IDFrecuencia = ?"
            params.append(int(idfrecuencia))
        else:
            where += " AND CAST(f.IDFrecuencia AS VARCHAR(50)) LIKE ?"
            params.append(f"%{idfrecuencia.strip()}%")
    if nombre:
        where += " AND f.Frecuencia LIKE ?"
        params.append(f"%{nombre.strip()}%")
    if situacion_id not in (None, ""):
        where += " AND f.PKIDSituacionRegistro = ?"
        params.append(int(situacion_id))
    return where

# --------------------- list ---------------------

@router.get("/")
def list_frecuencia(
    idfrecuencia: str | None = Query(default=None, description="IDFrecuencia (exacto o contiene)"),
    nombre: str | None = Query(default=None, description="Frecuencia (contiene)"),
    situacion_id: int | None = Query(default=None, description="PKIDSituacionRegistro"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
):
    try:
        conn = get_connection()
        cur = conn.cursor()

        params_total = []
        where_total = _build_where(params_total, idfrecuencia, nombre, situacion_id)
        cur.execute(f"""
            SELECT COUNT(1) AS total
            FROM Frecuencia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where_total}
        """, params_total)
        total = int(cur.fetchone().total)

        offset = (page - 1) * page_size
        params = []
        where = _build_where(params, idfrecuencia, nombre, situacion_id)
        cur.execute(f"""
            SELECT f.PKID, f.IDFrecuencia, f.Frecuencia, f.NumeroDias, f.NumeroHoras,
                   f.PKIDSituacionRegistro, sr.SituacionRegistro
            FROM Frecuencia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY f.IDFrecuencia ASC, f.Frecuencia ASC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, page_size])
        items = [_row_to_dict(r) for r in cur.fetchall()]
        return {"items": items, "total": total}
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- create ---------------------

@router.post("/")
def create_frecuencia(payload: dict):
    _validate_payload(payload)
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO Frecuencia
              (IDFrecuencia, Frecuencia, NumeroDias, NumeroHoras, PKIDSituacionRegistro)
            VALUES (?,?,?,?,?)
        """, (
            payload.get("IDFrecuencia"),
            payload["Frecuencia"],
            payload["NumeroDias"],
            payload["NumeroHoras"],
            payload["PKIDSituacionRegistro"],
        ))
        conn.commit()
        return {"detail": "Creado"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad (único/FK).")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- update ---------------------

@router.put("/{pkid}")
def update_frecuencia(pkid: int, payload: dict):
    _validate_payload(payload)
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT 1 FROM Frecuencia WHERE PKID = ?", (pkid,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="No encontrado.")

        cur.execute("""
            UPDATE Frecuencia
               SET IDFrecuencia = ?,
                   Frecuencia = ?,
                   NumeroDias = ?,
                   NumeroHoras = ?,
                   PKIDSituacionRegistro = ?
             WHERE PKID = ?
        """, (
            payload.get("IDFrecuencia"),
            payload["Frecuencia"],
            payload["NumeroDias"],
            payload["NumeroHoras"],
            payload["PKIDSituacionRegistro"],
            pkid,
        ))
        conn.commit()
        return {"detail": "Actualizado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad (único/FK).")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- delete ---------------------

@router.delete("/{pkid}")
def delete_frecuencia(pkid: int):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM Frecuencia WHERE PKID = ?", (pkid,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado.")
        conn.commit()
        return {"detail": "Eliminado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="No se puede eliminar: hay referencias.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- export CSV ---------------------

@router.get("/export")
def export_csv(
    idfrecuencia: str | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        where = _build_where(params, idfrecuencia, nombre, situacion_id)
        cur.execute(f"""
            SELECT f.IDFrecuencia, f.Frecuencia, f.NumeroDias, f.NumeroHoras, sr.SituacionRegistro
            FROM Frecuencia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY f.IDFrecuencia ASC, f.Frecuencia ASC
        """, params)
        rows = cur.fetchall()

        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(["IDFrecuencia", "Frecuencia", "NumeroDias", "NumeroHoras", "Situacion"])
        for r in rows:
            w.writerow([r.IDFrecuencia, r.Frecuencia, r.NumeroDias, r.NumeroHoras, r.SituacionRegistro or ""])

        out = io.BytesIO(si.getvalue().encode("utf-8-sig"))
        headers = {"Content-Disposition": 'attachment; filename="frecuencia.csv"'}
        return StreamingResponse(out, media_type="text/csv", headers=headers)
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- export XLSX ---------------------

@router.get("/export-xlsx")
def export_xlsx(
    idfrecuencia: str | None = Query(default=None),
    nombre: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        where = _build_where(params, idfrecuencia, nombre, situacion_id)
        cur.execute(f"""
            SELECT f.IDFrecuencia, f.Frecuencia, f.NumeroDias, f.NumeroHoras, sr.SituacionRegistro
            FROM Frecuencia f
            LEFT JOIN SituacionRegistro sr ON f.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY f.IDFrecuencia ASC, f.Frecuencia ASC
        """, params)
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Frecuencia"
        headers = ["IDFrecuencia", "Frecuencia", "NumeroDias", "NumeroHoras", "Situación"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for r in rows:
            ws.append([r.IDFrecuencia, r.Frecuencia, r.NumeroDias, r.NumeroHoras, r.SituacionRegistro or ""])

        # auto-width
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
        ws.freeze_panes = "A2"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="frecuencia.xlsx"'}
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
