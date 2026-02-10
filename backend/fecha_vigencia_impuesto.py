# archivo: fecha_vigencia_impuesto.py
from fastapi import APIRouter, Depends, HTTPException, Query
from starlette.responses import StreamingResponse
import pyodbc
import io
import csv
from datetime import date

from database import get_connection
from security import get_current_user

# XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(
    prefix="/fecha-vigencia-impuesto",
    tags=["FechaVigenciaImpuesto"],
    dependencies=[Depends(get_current_user)],
)

# --------------------- helpers ---------------------

def _row_to_dict(r):
    return {
        "PKID": r.PKID,
        "IDTipoImpuesto": r.IDTipoImpuesto,
        "FechaVigencia": r.FechaVigencia.isoformat() if r.FechaVigencia else None,
        "TasaImpuesto": float(r.TasaImpuesto) if r.TasaImpuesto is not None else None,
        "ImporteBase": float(r.ImporteBase) if r.ImporteBase is not None else None,
        "PKIDSituacionRegistro": r.PKIDSituacionRegistro,
        "SituacionRegistro": r.SituacionRegistro,
    }

def _build_where(params, idtipo: str | None, fdesde: str | None, fhasta: str | None, situacion_id: int | None):
    where = " WHERE 1=1"
    if idtipo:
        where += " AND fvi.IDTipoImpuesto LIKE ?"
        params.append(f"%{idtipo.strip()}%")
    if fdesde:
        where += " AND fvi.FechaVigencia >= ?"
        params.append(fdesde)
    if fhasta:
        where += " AND fvi.FechaVigencia <= ?"
        params.append(fhasta)
    if situacion_id not in (None, ""):
        where += " AND fvi.PKIDSituacionRegistro = ?"
        params.append(int(situacion_id))
    return where

def _validate_payload(p: dict, updating: bool = False):
    # IDTipoImpuesto: requerido (char(3))
    it = (p.get("IDTipoImpuesto") or "").strip().upper()
    if not it:
        raise HTTPException(status_code=422, detail="Falta campo requerido: IDTipoImpuesto")
    if len(it) > 3:
        raise HTTPException(status_code=422, detail="IDTipoImpuesto debe tener máximo 3 caracteres")
    p["IDTipoImpuesto"] = it

    # FechaVigencia: requerido (date)
    if not p.get("FechaVigencia"):
        raise HTTPException(status_code=422, detail="Falta campo requerido: FechaVigencia")

    # TasaImpuesto: requerido (decimal >= 0)
    try:
        ti = float(p.get("TasaImpuesto"))
    except Exception:
        raise HTTPException(status_code=422, detail="TasaImpuesto debe ser numérico")
    if ti < 0:
        raise HTTPException(status_code=422, detail="TasaImpuesto no puede ser negativo")
    p["TasaImpuesto"] = ti

    # ImporteBase: opcional (>= 0 si viene)
    if p.get("ImporteBase") not in (None, ""):
        try:
            ib = float(p.get("ImporteBase"))
        except Exception:
            raise HTTPException(status_code=422, detail="ImporteBase debe ser numérico")
        if ib < 0:
            raise HTTPException(status_code=422, detail="ImporteBase no puede ser negativo")
        p["ImporteBase"] = ib
    else:
        p["ImporteBase"] = None

    # SituacionRegistro: requerido
    if p.get("PKIDSituacionRegistro") in (None, ""):
        raise HTTPException(status_code=422, detail="Falta campo requerido: PKIDSituacionRegistro")
    try:
        p["PKIDSituacionRegistro"] = int(p["PKIDSituacionRegistro"])
    except Exception:
        raise HTTPException(status_code=422, detail="PKIDSituacionRegistro debe ser entero")

# --------------------- list ---------------------

@router.get("/")
def list_fvi(
    idtipo: str | None = Query(default=None, description="Filtra por IDTipoImpuesto (contiene)"),
    fecha_desde: str | None = Query(default=None, description="YYYY-MM-DD"),
    fecha_hasta: str | None = Query(default=None, description="YYYY-MM-DD"),
    situacion_id: int | None = Query(default=None, description="PKIDSituacionRegistro"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
):
    try:
        conn = get_connection()
        cur = conn.cursor()

        params_total = []
        where_total = _build_where(params_total, idtipo, fecha_desde, fecha_hasta, situacion_id)
        sql_total = f"""
            SELECT COUNT(1) AS total
            FROM FechaVigenciaImpuesto fvi
            LEFT JOIN SituacionRegistro sr ON fvi.PKIDSituacionRegistro = sr.PKID
            {where_total}
        """
        cur.execute(sql_total, params_total)
        total = int(cur.fetchone().total)

        offset = (page - 1) * page_size
        params = []
        where = _build_where(params, idtipo, fecha_desde, fecha_hasta, situacion_id)
        sql = f"""
            SELECT fvi.PKID, fvi.IDTipoImpuesto, fvi.FechaVigencia, fvi.TasaImpuesto,
                   fvi.ImporteBase, fvi.PKIDSituacionRegistro, sr.SituacionRegistro
            FROM FechaVigenciaImpuesto fvi
            LEFT JOIN SituacionRegistro sr ON fvi.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY fvi.FechaVigencia DESC, fvi.IDTipoImpuesto ASC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        params += [offset, page_size]
        cur.execute(sql, params)
        items = [_row_to_dict(r) for r in cur.fetchall()]
        return {"items": items, "total": total}
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- create ---------------------

@router.post("/")
def create_fvi(payload: dict):
    _validate_payload(payload, updating=False)
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO FechaVigenciaImpuesto
              (IDTipoImpuesto, FechaVigencia, TasaImpuesto, ImporteBase, PKIDSituacionRegistro)
            VALUES (?,?,?,?,?)
        """, (
            payload["IDTipoImpuesto"],
            payload["FechaVigencia"],
            payload["TasaImpuesto"],
            payload["ImporteBase"],
            payload["PKIDSituacionRegistro"],
        ))
        conn.commit()
        return {"detail": "Creado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad / FK.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- update ---------------------

@router.put("/{pkid}")
def update_fvi(pkid: int, payload: dict):
    _validate_payload(payload, updating=True)
    try:
        conn = get_connection()
        cur = conn.cursor()

        cur.execute("SELECT 1 FROM FechaVigenciaImpuesto WHERE PKID = ?", (pkid,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="No encontrado.")

        cur.execute("""
            UPDATE FechaVigenciaImpuesto
               SET IDTipoImpuesto = ?,
                   FechaVigencia = ?,
                   TasaImpuesto = ?,
                   ImporteBase = ?,
                   PKIDSituacionRegistro = ?
             WHERE PKID = ?
        """, (
            payload["IDTipoImpuesto"],
            payload["FechaVigencia"],
            payload["TasaImpuesto"],
            payload["ImporteBase"],
            payload["PKIDSituacionRegistro"],
            pkid,
        ))
        conn.commit()
        return {"detail": "Actualizado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="Violación de integridad / FK.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- delete ---------------------

@router.delete("/{pkid}")
def delete_fvi(pkid: int):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM FechaVigenciaImpuesto WHERE PKID = ?", (pkid,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado.")
        conn.commit()
        return {"detail": "Eliminado"}
    except pyodbc.IntegrityError:
        conn.rollback()
        raise HTTPException(status_code=409, detail="No se puede eliminar: hay referencias.")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- export CSV ---------------------

@router.get("/export")
def export_csv(
    idtipo: str | None = Query(default=None),
    fecha_desde: str | None = Query(default=None),
    fecha_hasta: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        where = _build_where(params, idtipo, fecha_desde, fecha_hasta, situacion_id)
        sql = f"""
            SELECT fvi.IDTipoImpuesto, fvi.FechaVigencia, fvi.TasaImpuesto,
                   fvi.ImporteBase, sr.SituacionRegistro
            FROM FechaVigenciaImpuesto fvi
            LEFT JOIN SituacionRegistro sr ON fvi.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY fvi.FechaVigencia DESC, fvi.IDTipoImpuesto ASC
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(["IDTipoImpuesto", "FechaVigencia", "TasaImpuesto", "ImporteBase", "Situacion"])
        for r in rows:
            w.writerow([
                r.IDTipoImpuesto,
                r.FechaVigencia.strftime("%Y-%m-%d") if r.FechaVigencia else "",
                float(r.TasaImpuesto) if r.TasaImpuesto is not None else "",
                float(r.ImporteBase) if r.ImporteBase is not None else "",
                r.SituacionRegistro or "",
            ])

        out = io.BytesIO(si.getvalue().encode("utf-8-sig"))
        headers = {"Content-Disposition": 'attachment; filename="fecha_vigencia_impuesto.csv"'}
        return StreamingResponse(out, media_type="text/csv", headers=headers)
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------- export XLSX ---------------------

@router.get("/export-xlsx")
def export_xlsx(
    idtipo: str | None = Query(default=None),
    fecha_desde: str | None = Query(default=None),
    fecha_hasta: str | None = Query(default=None),
    situacion_id: int | None = Query(default=None),
):
    try:
        conn = get_connection()
        cur = conn.cursor()
        params = []
        where = _build_where(params, idtipo, fecha_desde, fecha_hasta, situacion_id)
        sql = f"""
            SELECT fvi.IDTipoImpuesto, fvi.FechaVigencia, fvi.TasaImpuesto,
                   fvi.ImporteBase, sr.SituacionRegistro
            FROM FechaVigenciaImpuesto fvi
            LEFT JOIN SituacionRegistro sr ON fvi.PKIDSituacionRegistro = sr.PKID
            {where}
            ORDER BY fvi.FechaVigencia DESC, fvi.IDTipoImpuesto ASC
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "FechaVigencia"
        headers = ["IDTipoImpuesto", "FechaVigencia", "TasaImpuesto", "ImporteBase", "Situación"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for r in rows:
            ws.append([
                r.IDTipoImpuesto,
                r.FechaVigencia.strftime("%Y-%m-%d") if r.FechaVigencia else "",
                float(r.TasaImpuesto) if r.TasaImpuesto is not None else "",
                float(r.ImporteBase) if r.ImporteBase is not None else "",
                r.SituacionRegistro or "",
            ])

        # auto-width
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
        ws.freeze_panes = "A2"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="fecha_vigencia_impuesto.xlsx"'}
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
